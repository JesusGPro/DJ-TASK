from django.shortcuts import render, redirect, get_object_or_404
from tasks.models import ActiveProject, Project
from prices.models import Task, Price
from workpackage.models import WorkPackage, WorkpackageTotals, Work, Measurement
from workpackage.forms import WPIdForm
from django.http import HttpResponse
from weasyprint import HTML
from django.template.loader import render_to_string
import csv
from decimal import Decimal
# import pandas as pd
from django.contrib import messages
from django.db.models import Sum
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

current_datetime = datetime.now().strftime("%Y-%m-%d_%H-%M")  # Format as YYYY-MM-DD_HH-MM

def work_package_assign(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description
    wps = WorkPackage.objects.all().order_by('name')

    if request.method == 'POST':
        form = WPIdForm(request.POST)
        work_package_id = request.POST.get('work_package_id')
        if 'view_report' in request.POST and work_package_id:
            return redirect('work_package_report', work_package_id)
        elif 'pdf_button' in request.POST and work_package_id:
            return redirect('work_package_pdf', work_package_id)
        elif 'csv_button' in request.POST and work_package_id:
            return redirect('work_package_csv', work_package_id)
        elif 'view_measurements' in request.POST and work_package_id:
            return redirect('work_package_report_measurement', work_package_id)
        elif 'view_measurements_pdf' in request.POST and work_package_id:
            return redirect('work_package_report_measurement_pdf', work_package_id)
        elif 'view_measurements_csv' in request.POST and work_package_id:
            return redirect('work_package_report_measurement_csv', work_package_id)
        elif 'view_measurements_xlsx' in request.POST and work_package_id:
            return redirect('work_package_report_measurement_xlsx', work_package_id)
        else:
            messages.error(request,"You have to choose a WorkPackage")
    else:
        form = WPIdForm()
   
    return render(request, 'reports/work_package_assign.html', {
        'form': form,
        'project_id': project_id, 
        'project_description': project_description,
        'wps': wps
        })


def work_package_report(request, work_package_id):
    work_package_id = work_package_id
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_package = WorkPackage.objects.get(id=work_package_id)
    tasks = Task.objects.all()

    # Sending elements to the list
    works = Work.objects.filter(work_package=work_package_id).select_related('task')
    task_ids = list(Work.objects.filter(work_package=work_package_id).values_list('task_id', flat=False))
    task_ids = [task_id[0] for task_id in task_ids]
    tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
    zips = list(zip(works, tasks_retrieved))

    grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package_id)
    grand_total = float(grand_total.total)

    formatted_grand_total = "{:,.2f}".format(grand_total)

    context = {
        'project_id': project_id,
        'project_description': project_description,
        'work_package_id': work_package_id,
        'work_package_name': work_package.name,
        'work_package_description': work_package.description,
        'tasks': tasks,
        'works': works,
        'tasks_retrieved': tasks_retrieved,
        'zips': zips,
        'formatted_grand_total': formatted_grand_total
    }

    return render(request, 'reports/work_package_report.html', context)

def work_package_pdf(request, work_package_id):
    work_package_id = work_package_id
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_package = WorkPackage.objects.get(id=work_package_id)
    tasks = Task.objects.all()

    # Sending elements to the list
    works = Work.objects.filter(work_package=work_package_id).select_related('task')
    task_ids = list(Work.objects.filter(work_package=work_package_id).values_list('task_id', flat=False))
    task_ids = [task_id[0] for task_id in task_ids]
    tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
    zips = list(zip(works, tasks_retrieved))

    grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package_id)
    grand_total = float(grand_total.total)

    formatted_grand_total = "{:,.2f}".format(grand_total)

    # Creating the pdf
    html_template = 'reports/work_package_report.html'
    context = {
        'project_id': project_id,
        'project_description': project_description,
        'work_package_id': work_package_id,
        'work_package_name': work_package.name,
        'work_package_description': work_package.description,
        'tasks': tasks,
        'works': works,
        'tasks_retrieved': tasks_retrieved,
        'zips': zips,
        'formatted_grand_total': formatted_grand_total
    }
    
    html_string = render_to_string(html_template, context)

    pdf_file = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_{current_datetime}.pdf"'

    return response



def work_package_csv(request, work_package_id):
    work_package_id = work_package_id
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_package = WorkPackage.objects.get(id=work_package_id)
    tasks = Task.objects.all()

    # Sending elements to the list
    works = Work.objects.filter(work_package=work_package_id).select_related('task')
    task_ids = list(Work.objects.filter(work_package=work_package_id).values_list('task_id', flat=False))
    task_ids = [task_id[0] for task_id in task_ids]
    tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
    zips = list(zip(works, tasks_retrieved))

    grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package_id)
    grand_total = float(grand_total.total)

    formatted_grand_total = "{:,.2f}".format(grand_total)

    # Creating the CSV file
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_{current_datetime}.csv"'

    writer = csv.writer(response)

    writer.writerow([
        'Project ID',
        'Project Description',
        'Work Package ID',
        'Work Package Name',
        'Work Package Description',
        'Grand Total'
        
    ])
    writer.writerow([
        project_id,
        project_description,
        work_package_id,
        work_package.name,
        work_package.description,
        formatted_grand_total,
        None
    ])
    
    writer.writerow([
        ' ',
        ' ',
    ])

    writer.writerow([
        'Task code',
        'Task name',
        'Unit',
        'Price',
        'Quantity',
        'Amount'
    ])

    for work, task in zips:
        writer.writerow([
            task.code,
            task.name,
            task.unit,
            task.price,
            format(work.quantity, 'g'),
            format(work.work_amount, 'g')
        ])

    return response
######################### WorkPackage Report Total ###########################################################3
def work_package_total_pick(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    if request.method == 'POST':
        form = WPIdForm(request.POST)
        
        if 'view_report' in request.POST:
            return redirect('work_package_report_total')
        elif 'pdf_button' in request.POST:
            return redirect('work_package_report_total_pdf')
        elif 'csv_button' in request.POST:
            return redirect('work_package_report_total_csv')
        elif 'view_measurements' in request.POST:
            return redirect('work_package_report_measurement_total')
        elif 'view_measurements_pdf' in request.POST:
            return redirect('work_package_report_measurement_total_pdf')
        elif 'view_measurements_csv' in request.POST:
            return redirect('work_package_report_measurement_total_csv')
        elif 'view_measurements_xlsx' in request.POST:
            return redirect('work_package_report_measurement_total_excel')
        elif 'view_report_summary' in request.POST:
            return redirect('work_package_report_total_summary')
        elif 'view_report_summary_pdf' in request.POST:
            return redirect('work_package_report_total_summary_pdf')
        elif 'view_report_summary_csv' in request.POST:
            return redirect('work_package_report_total_summary_csv')
        else:
            messages.error(request,"This option is not yet available")
    else:
        form = WPIdForm()
   
    return render(request, 'reports/work_package_report_total_index.html', {
        'form': form,
        'project_id': project_id, 
        'project_description': project_description
        })

def calculate_chapter_totals(wps):
    work_packages = defaultdict(list)

    for wp in wps:       
        # Ensure the work package itself is in the dictionary
        work_packages[wp.name]  # This initializes the entry for every work package
        
        if wp.parent_id:
            work_packages[wp.parent.name].append(wp.name)


    work_package_totals = {}

    for wp in wps:
        try:
            work_package_totals[wp.name] = WorkpackageTotals.objects.get(workpackage_id=wp.id).total 

        except WorkpackageTotals.DoesNotExist:
            # If the total does not exist, set it to 0
            work_package_totals[wp.name] = 0
    

    # Function to calculate totals recursively
    def calculate_total(key):
        total = Decimal('0.00')
        # Check if the work package has child packages
        if key in work_packages:
            for child in work_packages[key]:
                total += calculate_total(child)  # Recursively sum child totals
        # Add the total for the current work package if it exists
        if key in work_package_totals:
            total += work_package_totals[key]
        return total

    # Calculate totals for each chapter
    chapter_totals = {}
    for key in work_packages.keys():
        chapter_totals[key] = calculate_total(key)


    # Format the accumulated totals for display
    formatted_totals = {key: "{:,.2f}".format(value) for key, value in chapter_totals.items()}
    
    return formatted_totals

def work_package_report_total(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_packages = WorkPackage.objects.all().order_by('name')
    
    # Prepare the data for each work package
    reports = []
    
    # Get the total amount for the project directly from WorkpackageTotals
    project_total = WorkpackageTotals.objects.filter(workpackage__in=work_packages).aggregate(total=Sum('total'))['total'] or 0.0
    formatted_project_total = "{:,.2f}".format(project_total)

    for work_package in work_packages:
        works = Work.objects.filter(work_package=work_package.id).select_related('task')

        zips= []
        
        # Check if there are any works for the current work package
        if works.exists():
            # Extract task IDs from the works
            task_ids = works.values_list('task_id', flat=True)
           
            # Retrieve tasks based on the task IDs found
            tasks_retrieved = Task.objects.filter(id__in=task_ids).order_by('id')  # You can order as needed
            # Create a mapping of task IDs to task objects for easy lookup
            task_dict = {task.id: task for task in tasks_retrieved}
            # Prepare the zipped data
            zips = [(work, task_dict.get(work.task_id)) for work in works]
            
            # Attempt to get the grand total, handle the case where it doesn't exist
            try:
                grand_total = WorkpackageTotals.objects.get(workpackage=work_package)
                grand_total_value = float(grand_total.total)
                formatted_grand_total = "{:,.2f}".format(grand_total_value)
            except WorkpackageTotals.DoesNotExist:
                formatted_grand_total = "0.00"  
        else:
            tasks_retrieved = []  # No works means no tasks
            formatted_grand_total = "0.00"  # or handle it as you prefer

        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'zips': zips,
            'formatted_grand_total': formatted_grand_total
        })

    # Calculate chapter totals using the separate function
    chapter_totals = calculate_chapter_totals(work_packages)

    context = {
        'project_id': project_id,
        'project_description': project_description,
        'reports': reports,
        'formatted_project_total': formatted_project_total,
        'chapter_totals': chapter_totals
    }

    return render(request, 'reports/work_package_report_total.html', context)


def work_package_report_total_pdf(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_packages = WorkPackage.objects.all().order_by('name')
    
    # Prepare the data for each work package
    reports = []
    
    # Get the total amount for the project directly from WorkpackageTotals
    project_total = WorkpackageTotals.objects.filter(workpackage__in=work_packages).aggregate(total=Sum('total'))['total'] or 0.0
    formatted_project_total = "{:,.2f}".format(project_total)

    for work_package in work_packages:
        works = Work.objects.filter(work_package=work_package.id).select_related('task')

        zips= []
        
        # Check if there are any works for the current work package
        if works.exists():
            # Extract task IDs from the works
            task_ids = works.values_list('task_id', flat=True)
           
            # Retrieve tasks based on the task IDs found
            tasks_retrieved = Task.objects.filter(id__in=task_ids).order_by('id')  # You can order as needed
            # Create a mapping of task IDs to task objects for easy lookup
            task_dict = {task.id: task for task in tasks_retrieved}
            # Prepare the zipped data
            zips = [(work, task_dict.get(work.task_id)) for work in works]
            
            # Attempt to get the grand total, handle the case where it doesn't exist
            try:
                grand_total = WorkpackageTotals.objects.get(workpackage=work_package)
                grand_total_value = float(grand_total.total)
                formatted_grand_total = "{:,.2f}".format(grand_total_value)
            except WorkpackageTotals.DoesNotExist:
                formatted_grand_total = "0.00"  
        else:
            tasks_retrieved = []  # No works means no tasks
            formatted_grand_total = "0.00"  # or handle it as you prefer

        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'zips': zips,
            'formatted_grand_total': formatted_grand_total
        })

    # Calculate chapter totals using the separate function
    chapter_totals = calculate_chapter_totals(work_packages)

    context = {
        'project_id': project_id,
        'project_description': project_description,
        'reports': reports,
        'formatted_project_total': formatted_project_total,
        'chapter_totals': chapter_totals
    }

    # Render the HTML template
    html_string = render(request, 'reports/work_package_report_total.html', context).content

    # Convert HTML to PDF
    pdf = HTML(string=html_string).write_pdf()

    # Create the HTTP response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_total_{current_datetime}.pdf"'
    return response


def work_package_report_total_csv(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_packages = WorkPackage.objects.all().order_by('name')
    
    # Prepare the HTTP response for CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_total_{current_datetime}.csv"'

    writer = csv.writer(response)
    
    # Write the header
    writer.writerow(['Project Description', project_description])
    writer.writerow(['Project ID', project_id])
    writer.writerow([])
    
    # Write the reports header
    writer.writerow(['Work Package description', 'Tasks breakdown', 'Grand Totol for Work Package', ' ', 'Amount'])
    
    # Write each work package's details
    for work_package in work_packages:
        works = Work.objects.filter(work_package=work_package.id).select_related('task')
        grand_total_value = WorkpackageTotals.objects.filter(workpackage=work_package).aggregate(total=Sum('total'))['total'] or 0.0
        formatted_grand_total = "{:,.2f}".format(grand_total_value)
        
        # Get chapter total
        chapter_totals = calculate_chapter_totals(work_packages)
        chapter_total = chapter_totals.get(work_package.name, "0.00")
        
        # Write work package summary
        # writer.writerow([work_package.id, work_package.name, work_package.description, chapter_total])
        writer.writerow([work_package.description, chapter_total])
        
        # Write task details if exists
        if works.exists():
            writer.writerow(['Task Code', 'Task Name', 'Unit', 'Price', 'Quantity', 'Amount'])
            for work in works:
                task = work.task  # Assuming each work has a related task
                writer.writerow([
                    task.code,
                    task.name,
                    task.unit,
                    task.price,
                    work.quantity,
                    work.work_amount
                ])
        
        # Write grand total for the work package
        writer.writerow(['', '', 'Grand Total for Work Package', '', '', formatted_grand_total])
        writer.writerow([])  # Add a blank line between work packages

    return response

######################### WorkPackage Report Total summary ############################################################33333333333
def work_package_report_total_summary(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_packages = WorkPackage.objects.all().order_by('name')
    # Calculate chapter totals using the separate function
    chapter_totals = calculate_chapter_totals(work_packages)

    
    # Prepare the data for each work package
    reports = []
    
    # # Get the total amount for the project directly from WorkpackageTotals
    project_total = WorkpackageTotals.objects.filter(workpackage__in=work_packages).aggregate(total=Sum('total'))['total'] or 0.0
    formatted_project_total = "{:,.2f}".format(project_total)

    for work_package in work_packages:
        # Calculate the grand total for the current work package
        grand_total = chapter_totals[work_package.name]
        
        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'formatted_grand_total': grand_total
        })


    context = {
        'project_id': project_id,
        'project_description': project_description,
        'reports': reports,
        'formatted_project_total': formatted_project_total,
        'chapter_totals': chapter_totals
    }

    return render(request, 'reports/work_package_report_total_summary.html', context)

def work_package_report_total_summary_pdf(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_packages = WorkPackage.objects.all().order_by('name')
    # Calculate chapter totals using the separate function
    chapter_totals = calculate_chapter_totals(work_packages)
    
    # Prepare the data for each work package
    reports = []
    
    # # Get the total amount for the project directly from WorkpackageTotals
    project_total = WorkpackageTotals.objects.filter(workpackage__in=work_packages).aggregate(total=Sum('total'))['total'] or 0.0
    formatted_project_total = "{:,.2f}".format(project_total)

    for work_package in work_packages:
        # Calculate the grand total for the current work package
        grand_total = chapter_totals[work_package.name]
        
        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'formatted_grand_total': grand_total
        })


    context = {
        'project_id': project_id,
        'project_description': project_description,
        'reports': reports,
        'formatted_project_total': formatted_project_total,
        'chapter_totals': chapter_totals
    }
    # Render the HTML template
    html_string = render_to_string('reports/work_package_report_total_summary.html', context)
    
    # Create a PDF from the HTML
    pdf = HTML(string=html_string).write_pdf()

    # Return the PDF as a response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_summary_{current_datetime}.pdf"'
    return response

def work_package_report_total_summary_csv(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_packages = WorkPackage.objects.all().order_by('name')
    # Calculate chapter totals using the separate function
    chapter_totals = calculate_chapter_totals(work_packages)
    
    # Prepare the data for each work package
    reports = []
    
    # # Get the total amount for the project directly from WorkpackageTotals
    project_total = WorkpackageTotals.objects.filter(workpackage__in=work_packages).aggregate(total=Sum('total'))['total'] or 0.0
    formatted_project_total = "{:,.2f}".format(project_total)

    for work_package in work_packages:
        # Calculate the grand total for the current work package
        grand_total = chapter_totals[work_package.name]
        
        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'formatted_grand_total': grand_total
        })


    # Prepare the CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_summary_{current_datetime}.csv"'

    writer = csv.writer(response)
    # Write the header
    writer.writerow(['Project Description', project_description])
    writer.writerow(['Description', 'Grand Total'])

    # Write the data rows
    for report in reports:
        writer.writerow([
            report['work_package_description'],
            report['formatted_grand_total']
        ])

    # Optionally, you can add a summary row for the project total
    writer.writerow([])
    writer.writerow(['Project Total', formatted_project_total])

    return response

######################### WorkPackage Report Total with measurements ###########################################################3
def work_package_report_measurement_total(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    # Get all work packages for the active project
    work_packages = WorkPackage.objects.all().order_by('name')
    
    # Prepare context data
    reports = []
    
    for work_package in work_packages:
        # Retrieve works associated with the work package
        works = Work.objects.filter(work_package=work_package.id).select_related('task')
        
        # Retrieve task IDs and tasks
        task_ids = list(works.values_list('task_id', flat=True))
        tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
        
        # Create a zip of works and tasks
        zips = list(zip(works, tasks_retrieved))
        
        # Get the grand total for the work package
        try:
            grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package.id)
            grand_total_value = float(grand_total.total)
            formatted_grand_total = "{:,.2f}".format(grand_total_value)
        except WorkpackageTotals.DoesNotExist:
            formatted_grand_total = "0.00"
        
        # Collect measurements for all works in this work package
        measurements = []
        for work in works:
            measurements_for_work = Measurement.objects.filter(work_id=work.id)
            measurements.append({
                'work': work,
                'measurements': measurements_for_work
            })

        # Append report data for this work package
        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'zips': zips,
            'formatted_grand_total': formatted_grand_total,
            'measurements': measurements,
        })

    context = {
        'project_id': project_id,
        'project_description': project_description,
        'reports': reports,
    }

    return render(request, 'reports/work_package_report_measurements_total.html', context)

def work_package_report_measurement_total_pdf(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    # Get all work packages for the active project
    work_packages = WorkPackage.objects.all().order_by('name')
    
    reports = []
    
    for work_package in work_packages:
        works = Work.objects.filter(work_package=work_package.id).select_related('task')
        task_ids = list(works.values_list('task_id', flat=True))
        tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
        zips = list(zip(works, tasks_retrieved))

        try:
            grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package.id)
            formatted_grand_total = "{:,.2f}".format(float(grand_total.total))
        except WorkpackageTotals.DoesNotExist:
            formatted_grand_total = "0.00"

        measurements = []
        for work in works:
            measurements_for_work = Measurement.objects.filter(work_id=work.id)
            measurements.append({
                'work': work,
                'measurements': measurements_for_work
            })

        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'zips': zips,
            'formatted_grand_total': formatted_grand_total,
            'measurements': measurements,
        })

    context = {
        'project_id': project_id,
        'project_description': project_description,
        'reports': reports,
    }

    # Render the template to a string
    html_string = render(request, 'reports/work_package_report_measurements_total.html', context).content.decode('utf-8')

    # Create a PDF from the HTML
    pdf_file = HTML(string=html_string).write_pdf()

    # Create the response
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="work_package_measurement_report_{current_datetime}.pdf"'

    return response


def work_package_report_measurement_total_csv(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    # Get all work packages for the active project
    # work_packages = WorkPackage.objects.all()
    work_packages = WorkPackage.objects.all().order_by('name')
    
    # Prepare context data
    reports = []
    
    for work_package in work_packages:
        # Retrieve works associated with the work package
        works = Work.objects.filter(work_package=work_package.id).select_related('task')
        
        # Retrieve task IDs and tasks
        task_ids = list(works.values_list('task_id', flat=True))
        # tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
        tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: x.name)
        
        # Create a zip of works and tasks
        zips = list(zip(works, tasks_retrieved))
        
        # Get the grand total for the work package
        try:
            grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package.id)
            grand_total_value = float(grand_total.total)
            formatted_grand_total = "{:,.2f}".format(grand_total_value)
        except WorkpackageTotals.DoesNotExist:
            formatted_grand_total = "0.00"
        
        # Collect measurements for all works in this work package
        measurements = []
        for work in works:
            measurements_for_work = Measurement.objects.filter(work_id=work.id)
            measurements.append({
                'work': work,
                'measurements': measurements_for_work
            })

        # Append report data for this work package
        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'zips': zips,
            'formatted_grand_total': formatted_grand_total,
            'measurements': measurements,
        })

    # Create the HTTP response for the CSV file
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="work_package_measurement_report_total_{current_datetime}.csv"'

    writer = csv.writer(response)
    
    # Write the header row for the work packages
    writer.writerow(['Project Description', project_description])
    writer.writerow(['Work Package Description', 'Grand Total'])
    
    for report in reports:
        # Write the work package details
        writer.writerow([report['work_package_description'], report['formatted_grand_total']])
        
        # Write a header for tasks and measurements
        writer.writerow(['Task Code', 'Task Name', 'Unit', 'Price', 'Quantity', 'Amount'])
        
        # Write task details
        for work, task in report['zips']:
            writer.writerow([task.code, task.name, task.unit, task.price, work.quantity, work.work_amount])
            
            # Write a header for measurements
            writer.writerow(['Description', 'Nr', 'Width', 'Length', 'Height', 'Partial', 'Comment'])
            
            # Write measurement details
            for measurement in report['measurements']:
                for measurement_detail in measurement['measurements']:
                    writer.writerow([
                        measurement_detail.description,
                        measurement_detail.nr,
                        measurement_detail.width,
                        measurement_detail.length,
                        measurement_detail.height,
                        measurement_detail.partial,
                        measurement_detail.comment
                    ])
            # Add a blank row for separation between tasks
            writer.writerow([])

    return response


def work_package_report_measurement_total_excel(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    # Get all work packages for the active project
    # work_packages = WorkPackage.objects.all()
    work_packages = WorkPackage.objects.all().order_by('name')
    
    # Prepare context data
    reports = []
    
    for work_package in work_packages:
        # Retrieve works associated with the work package
        works = Work.objects.filter(work_package=work_package.id).select_related('task')
        
        # Retrieve task IDs and tasks
        task_ids = list(works.values_list('task_id', flat=True))
        # tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
        tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: x.name)
        
        # Create a zip of works and tasks
        zips = list(zip(works, tasks_retrieved))
        
        # Get the grand total for the work package
        try:
            grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package.id)
            grand_total_value = float(grand_total.total)
            formatted_grand_total = "{:,.2f}".format(grand_total_value)
        except WorkpackageTotals.DoesNotExist:
            formatted_grand_total = "0.00"
        
        # Collect measurements for all works in this work package
        measurements = []
        for work in works:
            measurements_for_work = Measurement.objects.filter(work_id=work.id)
            measurements.append({
                'work': work,
                'measurements': measurements_for_work
            })

        # Append report data for this work package
        reports.append({
            'work_package_id': work_package.id,
            'work_package_name': work_package.name,
            'work_package_description': work_package.description,
            'zips': zips,
            'formatted_grand_total': formatted_grand_total,
            'measurements': measurements,
        })

    # Create an Excel workbook and a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Work Package Report'

    # Create a bold font style
    bold_font = Font(bold=True)

    # Write the header row for the project description
    worksheet.append(['Project Description', project_description])
    worksheet.append(['Work Package Description', 'Grand Total'])


    for report in reports:
        # Write the work package details
        worksheet.append([report['work_package_description'], report['formatted_grand_total']])

        # Apply bold font to the work package details if there is at least one row
        if worksheet.max_row > 0:  # Check if there are rows in the worksheet
            for cell in worksheet[worksheet.max_row]:  # Get the last row
                cell.font = bold_font  # Apply bold font to each cell in the last row
        
        # Write a header for tasks and measurements
        worksheet.append(['Task Code', 'Task Name', 'Unit', 'Price', 'Quantity', 'Amount'])
        
        # Write task details
        for work, task in report['zips']:
            worksheet.append([task.code, task.name, task.unit, task.price, work.quantity, work.work_amount])
            
            # Write a header for measurements
            worksheet.append(['Description', 'Nr', 'Width', 'Length', 'Height', 'Partial', 'Comment'])
            
            # Apply bold font to the work package details if there is at least one row
            if worksheet.max_row > 0:  # Check if there are rows in the worksheet
                for cell in worksheet[worksheet.max_row]:  # Get the last row
                    cell.font = bold_font  # Apply bold font to each cell in the last row

            # Write measurement details
            for measurement in report['measurements']:
                for measurement_detail in measurement['measurements']:
                    worksheet.append([
                        measurement_detail.description,
                        measurement_detail.nr,
                        measurement_detail.width,
                        measurement_detail.length,
                        measurement_detail.height,
                        measurement_detail.partial,
                        measurement_detail.comment
                    ])
            # Add a blank row for separation between tasks
            worksheet.append([])

    # Create an HTTP response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="work_package_measurement_report_total_{current_datetime}.xlsx"'

    # Save the workbook to the response
    workbook.save(response)

    return response    


######################### Component ###########################################################################

def component_report(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    prices = Price.objects.all()

    return render(request, 'reports/component_report.html', {'project': project, 'prices': prices})

def component_report_pdf(request):
    # sending the info to the html:
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    prices = Price.objects.all()

    # Render the HTML template with data
    html_string = render_to_string('reports/component_report.html', {'project': project, 'prices': prices})

    # Create a PDF file from the HTML string
    pdf_file = HTML(string=html_string).write_pdf(pagesize='A4')

    # Create a response with the PDF file
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Component Report_{current_datetime}.pdf"'

    return response

def component_report_csv(request):
    # sending the info to the html:
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    prices = Price.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="components_{current_datetime}.csv"'

    writer = csv.writer(response)

    writer.writerow([
        'Project ID',
        'Project Description',        
    ])
    writer.writerow([
        project.id,
        project.description,
    ])
    
    writer.writerow([
        ' ',
        ' ',
    ])

    writer.writerow(['id', 'Code', 'Denomination', 'Unit', 'Price', 'Currency', 'Tag', 'Reference'])

    for price in prices:
        writer.writerow([price.id, price.code, price.denomination, price.unit, price.price, price.currency, price.tag, price.reference])

    return response

################################### Tasks ###############################################################3

def task_report(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    tasks = Task.objects.all()

    return render(request, 'reports/task_report.html', {'project': project, 'tasks': tasks})

def task_report_pdf(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    tasks = Task.objects.all()

    # Render the HTML template with data
    html_string = render_to_string('reports/task_report.html', {'project': project, 'tasks': tasks})

    # Create a PDF file from the HTML string
    pdf_file = HTML(string=html_string).write_pdf(pagesize='A4')

    # Create a response with the PDF file
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Task Report_{current_datetime}.pdf"'

    return response

def task_report_csv(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    tasks = Task.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="tasks_{current_datetime}.csv"'

    writer = csv.writer(response)

    writer.writerow([
        'Project ID',
        'Project Description',        
    ])
    writer.writerow([
        project.id,
        project.description,
    ])
    
    writer.writerow([
        ' ',
        ' ',
    ])

    writer.writerow(['id', 'Code', 'Denomination', 'Unit', 'Price', 'Currency'])

    for task in tasks:
        writer.writerow([task.id, task.code, task.name, task.unit, task.price, task.currency])

    return response


def task_report_weight(request):
    tasks_dict={}
    tasks_weight = {}
    works = Work.objects.all()
    tasks = Task.objects.all()
    for work in works:
        if work.task_id not in tasks_dict:
            tasks_dict[work.task_id] = Decimal(work.quantity) * Decimal(Task.objects.get(id=work.task_id).price)
        else:
            tasks_dict[work.task_id] += Decimal(work.quantity) * Decimal(Task.objects.get(id=work.task_id).price)
    
    sorted_dict = dict(sorted(tasks_dict.items(), key=lambda item: item[1], reverse=True))

    for task_id in sorted_dict:
        tasks_weight[task_id] = [
            Task.objects.get(id=task_id).code, 
            Task.objects.get(id=task_id).name, 
            Task.objects.get(id=task_id).unit,
            Task.objects.get(id=task_id).price,
            Task.objects.get(id=task_id).currency,
            tasks_dict[task_id]
            ]
   
    tasks = []

    for key, value in tasks_weight.items():
        tasks.append({
            'id': key,
            'code': value[0],
            'name': value[1],
            'unit': value[2],
            'price': value[3],
            'currency': value[4],
            'amount': "{:,.2f}".format(value[5])
        })
    
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)

    return render(request, 'reports/task_report_weight.html', {'tasks': tasks, 'project': project})


def task_report_weight_pdf(request):
    tasks_dict={}
    tasks_weight = {}
    works = Work.objects.all()
    tasks = Task.objects.all()
    for work in works:
        if work.task_id not in tasks_dict:
            tasks_dict[work.task_id] = Decimal(work.quantity) * Decimal(Task.objects.get(id=work.task_id).price)
        else:
            tasks_dict[work.task_id] += Decimal(work.quantity) * Decimal(Task.objects.get(id=work.task_id).price)
    
    sorted_dict = dict(sorted(tasks_dict.items(), key=lambda item: item[1], reverse=True))

    for task_id in sorted_dict:
        tasks_weight[task_id] = [
            Task.objects.get(id=task_id).code, 
            Task.objects.get(id=task_id).name, 
            Task.objects.get(id=task_id).unit,
            Task.objects.get(id=task_id).price,
            Task.objects.get(id=task_id).currency,
            tasks_dict[task_id]
            ]
   
    tasks = []

    for key, value in tasks_weight.items():
        tasks.append({
            'id': key,
            'code': value[0],
            'name': value[1],
            'unit': value[2],
            'price': value[3],
            'currency': value[4],
            'amount': "{:,.2f}".format(value[5])
        })
    
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)

    # Render the HTML template with data
    html_string = render_to_string('reports/task_report_weight.html', {'project': project, 'tasks': tasks})

    # Create a PDF file from the HTML string
    pdf_file = HTML(string=html_string).write_pdf(pagesize='A4')

    # Create a response with the PDF file
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Task Report Weighted_{current_datetime}.pdf"'

    return response


def task_report_weight_csv(request):
    tasks_dict={}
    tasks_weight = {}
    works = Work.objects.all()
    tasks = Task.objects.all()
    for work in works:
        if work.task_id not in tasks_dict:
            tasks_dict[work.task_id] = Decimal(work.quantity) * Decimal(Task.objects.get(id=work.task_id).price)
        else:
            tasks_dict[work.task_id] += Decimal(work.quantity) * Decimal(Task.objects.get(id=work.task_id).price)
    
    sorted_dict = dict(sorted(tasks_dict.items(), key=lambda item: item[1], reverse=True))

    for task_id in sorted_dict:
        tasks_weight[task_id] = [
            Task.objects.get(id=task_id).code, 
            Task.objects.get(id=task_id).name, 
            Task.objects.get(id=task_id).unit,
            Task.objects.get(id=task_id).price,
            Task.objects.get(id=task_id).currency,
            tasks_dict[task_id]
            ]
   
    tasks = []

    for key, value in tasks_weight.items():
        tasks.append({
            'id': key,
            'code': value[0],
            'name': value[1],
            'unit': value[2],
            'price': value[3],
            'currency': value[4],
            'amount': "{:,.2f}".format(value[5])
        })
    
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="tasks_weighted_{current_datetime}.csv"'

    writer = csv.writer(response)

    writer.writerow([
        'Project ID',
        'Project Description',        
    ])
    writer.writerow([
        project.id,
        project.description,
    ])
    
    writer.writerow([
        ' ',
        ' ',
    ])

    writer.writerow(['id', 'Code', 'Denomination', 'Unit', 'Price', 'Currency', 'Volume'])

    for task in tasks:
        writer.writerow([task['id'], task['code'], task['name'], task['unit'], task['price'], task['currency'], task['amount']])

    return response


def task_components_report(request):
    # Project id and name
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    # Tasks and components
    tasks = Task.objects.filter(project=project_id).prefetch_related('taskcomponent_set__code')
    """Task.objects.filter(project=project_id), retrieves a queryset of Task only for this project
    .prefetch_related('taskcomponent_set__code'), retrieves the related objects (components)
    we're telling Django to retrieve the TaskComponent objects related to each Task object, 
    and then retrieve the Price objects related to each TaskComponent object, all in a single database query.
    In the template, we can then access the prefetched related objects using the taskcomponent_set attribute 
    on each Task object, like this: task.taskcomponent_set.all. This will give us a QuerySet of 
    TaskComponent objects related to each Task object, and we can then access the Price objects related 
    to each TaskComponent object using the code attribute.
    """
    context = {
        # project data
        'project_id': project_id,
        'project_description': project_description,
        # task data
        'tasks': tasks,
    }

    return render(request, 'reports/tasks_components_report.html', context)

def task_components_report_pdf(request):
    # Project id and name
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    # Tasks and components
    tasks = Task.objects.filter(project=project_id).prefetch_related('taskcomponent_set__code')
    
    context = {
        # project data
        'project_id': project_id,
        'project_description': project_description,
        # task data
        'tasks': tasks,
    }

    # Render the HTML template with data
    html_string = render_to_string('reports/tasks_components_report.html', context)

    # Create a PDF file from the HTML string
    pdf_file = HTML(string=html_string).write_pdf(pagesize='A4')

    # Create a response with the PDF file
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Task_with_Components_Report_{current_datetime}.pdf"'

    return response

def task_components_report_csv(request):
    # Project id and name
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    # Tasks and components
    tasks = Task.objects.filter(project=project_id).prefetch_related('taskcomponent_set__code')
    
    context = {
        # project data
        'project_id': project_id,
        'project_description': project_description,
        # task data
        'tasks': tasks,
    }

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="Task_with_component_report_{current_datetime}.csv"'

    writer = csv.writer(response)

    writer.writerow([
        'Project ID',
        'Project Description',        
    ])
    writer.writerow([
        project_id,
        project_description,
    ])
    
    writer.writerow([
        ' ',
        ' ',
    ])

    for task in tasks:
        writer.writerow([
        'Id',
        'Code',
        'Task Name',
        'Unit',
        ' ',
        ' ',
        'Price',
        'Currency'
        ])

        writer.writerow([
            task.id,
            task.code,
            task.name,
            task.unit,
            '',
            '',
            task.price,
            task.currency
        ])
        writer.writerow([
            'Components for Task',
            '',
            '',
            '',
            '',
            '',
            ''
        ])
        writer.writerow([
            'Id',
            'Code',
            'Task Name',
            'Unit',
            'Quantity',
            'Price',
            'Subtotal',
            'Currency'
        ])
        for component in task.taskcomponent_set.all():
            subtotal = component.quantity * component.code.price
            subtotal_formated = "{:,.2f}".format(subtotal)
            
            writer.writerow([
                component.code.id,
                component.code.code,
                component.code.denomination,
                component.code.unit,
                component.quantity,
                component.code.price,
                subtotal_formated,
                component.code.currency
            ])

    return response



def task_components_report_excel(request):
    # Project id and name
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    # Tasks and components
    tasks = Task.objects.filter(project=project_id).prefetch_related('taskcomponent_set__code')

    # Create an Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Task with Component Report"

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15

    # Write project information
    header_font = Font(bold=True)
    ws.append(['Project ID', 'Project Description'])
    for cell in ws[1]:
        cell.font = header_font
    ws.append([project_id, project_description])
    ws.append([' ', ' '])

    # Write task information
    for task in tasks:
        ws.append(['Id', 'Code', 'Task Name', 'Unit', ' ', ' ', 'Price', 'Currency'])
        for cell in ws[ws.max_row]:
            cell.font = header_font
        ws.append([task.id, task.code, task.name, task.unit, '', '', task.price, task.currency])
        ws.append(['Components for Task', '', '', '', '', '', ''])
        ws.append(['Id', 'Code', 'Task Name', 'Unit', 'Quantity', 'Price', 'Subtotal', 'Currency'])
        for cell in ws[ws.max_row]:
            cell.font = header_font

        # Write component information
        for component in task.taskcomponent_set.all():
            subtotal = component.quantity * component.code.price
            subtotal_formated = "{:,.2f}".format(subtotal)

            ws.append([
                component.code.id,
                component.code.code,
                component.code.denomination,
                component.code.unit,
                component.quantity,
                component.code.price,
                subtotal_formated,
                component.code.currency
            ])

    # Set alignment for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    # Create the HttpResponse object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Task_with_component_report_{current_datetime}.xlsx"'

    # Save the Excel file to the HttpResponse object
    wb.save(response)

    return response







#################### Work Package with Measurements ##################################################
def work_package_report_measurement(request, work_package_id):
    work_package_id = work_package_id
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_package = WorkPackage.objects.get(id=work_package_id)
    tasks = Task.objects.all()

    # Sending elements to the list
    works = Work.objects.filter(work_package=work_package_id).select_related('task')
    task_ids = list(Work.objects.filter(work_package=work_package_id).values_list('task_id', flat=False))
    task_ids = [task_id[0] for task_id in task_ids]
    tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
    zips = list(zip(works, tasks_retrieved))

    grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package_id)
    grand_total = float(grand_total.total)

    formatted_grand_total = "{:,.2f}".format(grand_total)

    # Measurements. I do not know how to filter this
    measurements = []
    for work in works:
        measurements_for_work = Measurement.objects.filter(work_id=work.id)
        measurements.append(measurements_for_work)


    context = {
        'project_id': project_id,
        'project_description': project_description,
        'work_package_id': work_package_id,
        'work_package_name': work_package.name,
        'work_package_description': work_package.description,
        'tasks': tasks,
        'works': works,
        'tasks_retrieved': tasks_retrieved,
        'zips': zips,
        'formatted_grand_total': formatted_grand_total,
        'measurements': measurements,
    }

    return render(request, 'reports/work_package_report_measurements.html', context)


def work_package_report_measurement_pdf(request, work_package_id):
    work_package_id = work_package_id
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_package = WorkPackage.objects.get(id=work_package_id)
    tasks = Task.objects.all()

    # Sending elements to the list
    works = Work.objects.filter(work_package=work_package_id).select_related('task')
    task_ids = list(Work.objects.filter(work_package=work_package_id).values_list('task_id', flat=False))
    task_ids = [task_id[0] for task_id in task_ids]
    tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
    zips = list(zip(works, tasks_retrieved))

    grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package_id)
    grand_total = float(grand_total.total)

    formatted_grand_total = "{:,.2f}".format(grand_total)

    # Measurements. I do not know how to filter this
    measurements = []
    for work in works:
        measurements_for_work = Measurement.objects.filter(work_id=work.id)
        measurements.append(measurements_for_work)

    # Creating the pdf
    html_template = 'reports/work_package_report_measurements.html'
    
    context = {
        'project_id': project_id,
        'project_description': project_description,
        'work_package_id': work_package_id,
        'work_package_name': work_package.name,
        'work_package_description': work_package.description,
        'tasks': tasks,
        'works': works,
        'tasks_retrieved': tasks_retrieved,
        'zips': zips,
        'formatted_grand_total': formatted_grand_total,
        'measurements': measurements,
    }

    html_string = render_to_string(html_template, context)

    pdf_file = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_{current_datetime}.pdf"'

    return response


def work_package_report_measurement_csv(request, work_package_id):
    work_package_id = work_package_id
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_package = WorkPackage.objects.get(id=work_package_id)
    tasks = Task.objects.all()

    # Sending elements to the list
    works = Work.objects.filter(work_package=work_package_id).select_related('task')
    task_ids = list(Work.objects.filter(work_package=work_package_id).values_list('task_id', flat=False))
    task_ids = [task_id[0] for task_id in task_ids]
    tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
    zips = list(zip(works, tasks_retrieved))

    grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package_id)
    grand_total = float(grand_total.total)

    formatted_grand_total = "{:,.2f}".format(grand_total)

    # Measurements. I do not know how to filter this
    measurements = []
    for work in works:
        measurements_for_work = Measurement.objects.filter(work_id=work.id)
        measurements.append(measurements_for_work)
   
    # Creating the CSV file
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_measurements_{current_datetime}.csv"'

    writer = csv.writer(response)

    writer.writerow([
        'Project ID',
        'Project Description',
        'Work Package ID',
        'Work Package Name',
        'Work Package Description',
        'Grand Total'
        
    ])
    writer.writerow([
        project_id,
        project_description,
        work_package_id,
        work_package.name,
        work_package.description,
        formatted_grand_total,
        None
    ])
    
    writer.writerow([
        ' ',
        ' ',
    ])

    for i, (work, task) in enumerate(zips):
        writer.writerow([
        'Task code',
        'Task name',
        'Unit',
        'Price',
        'Quantity',
        'Amount'
        ])

        writer.writerow([
            task.code,
            task.name,
            task.unit,
            task.price,
            format(work.quantity, 'g'),
            format(work.work_amount, 'g')
        ])
        writer.writerow([
            'Measurements for Task',
            '',
            '',
            '',
            '',
            '',
            ''
        ])
        writer.writerow([
            'Description',
            'Nr',
            'Width',
            'Length',
            'Height',
            'Partial',
            'Comment'
        ])
        for measurement in measurements[i]:
            writer.writerow([
                measurement.description,
                measurement.nr,
                measurement.width,
                measurement.length,
                measurement.height,
                measurement.partial,
                measurement.comment
            ])

    return response


import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def work_package_report_measurement_xlsx(request, work_package_id):
    work_package_id = work_package_id
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    work_package = WorkPackage.objects.get(id=work_package_id)
    tasks = Task.objects.all()

    # Sending elements to the list
    works = Work.objects.filter(work_package=work_package_id).select_related('task')
    task_ids = list(Work.objects.filter(work_package=work_package_id).values_list('task_id', flat=False))
    task_ids = [task_id[0] for task_id in task_ids]
    tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
    zips = list(zip(works, tasks_retrieved))

    grand_total = WorkpackageTotals.objects.get(workpackage_id=work_package_id)
    grand_total = float(grand_total.total)

    formatted_grand_total = "{:,.2f}".format(grand_total)

    measurements = []
    for work in works:
        measurements_for_work = Measurement.objects.filter(work_id=work.id)
        measurements.append(measurements_for_work)

    # Create a new Excel file
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set the column headers with bold font
    headers = [
        'Project ID',
        'Project Description',
        'Work Package ID',
        'Work Package Name',
        'Work Package Description',
        'Grand Total'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)

    # Add data to the worksheet
    row = 2
    ws.cell(row=row, column=1).value = project_id
    ws.cell(row=row, column=2).value = project_description
    ws.cell(row=row, column=3).value = work_package_id
    ws.cell(row=row, column=4).value = work_package.name
    ws.cell(row=row, column=5).value = work_package.description
    ws.cell(row=row, column=6).value = formatted_grand_total

    row += 2
    for i, (work, task) in enumerate(zips):
        ws.cell(row=row, column=1).value = 'Task code'
        ws.cell(row=row, column=2).value = 'Task name'
        ws.cell(row=row, column=3).value = 'Unit'
        ws.cell(row=row, column=4).value = 'Price'
        ws.cell(row=row, column=5).value = 'Quantity'
        ws.cell(row=row, column=6).value = 'Amount'

        row += 1
        ws.cell(row=row, column=1).value = task.code
        ws.cell(row=row, column=2).value = task.name
        ws.cell(row=row, column=3).value = task.unit
        ws.cell(row=row, column=4).value = task.price
        ws.cell(row=row, column=5).value = format(work.quantity, 'g')
        ws.cell(row=row, column=6).value = format(work.work_amount, 'g')

        row += 1
        ws.cell(row=row, column=1).value = 'Measurements for Task'
        ws.cell(row=row, column=2).value = ''
        ws.cell(row=row, column=3).value = ''
        ws.cell(row=row, column=4).value = ''
        ws.cell(row=row, column=5).value = ''
        ws.cell(row=row, column=6).value = ''

        row += 1
        headers = [
            'Description',
            'Nr',
            'Width',
            'Length',
            'Height',
            'Partial',
            'Comment'
        ]

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True)

        row += 1
        for measurement in measurements[i]:
            ws.cell(row=row, column=1).value = measurement.description
            ws.cell(row=row, column=2).value = measurement.nr
            ws.cell(row=row, column=3).value = measurement.width
            ws.cell(row=row, column=4).value = measurement.length
            ws.cell(row=row, column=5).value = measurement.height
            ws.cell(row=row, column=6).value = measurement.partial
            ws.cell(row=row, column=7).value = measurement.comment
            row += 1

    # Save the file

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="workpackage_report_measurement_{current_datetime}.xlsx"'
    wb.save(response)

    return response


###################### Workpackage report with children ###################################
def work_packagechildren_assign(request):
    active_project = ActiveProject.objects.get(user=request.user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description
    wps = WorkPackage.objects.all().order_by('name')
    name = []
    for wp in wps:
        if len(wp.name) > 3:
            remaining_part = wp.name[3:]
            if remaining_part== '000' and wp.name[0] !='0' and wp.name[1] !='0'and wp.name[2] !='0':
                name.append(wp)

    wps = name
       
    if request.method == 'POST':
        form = WPIdForm(request.POST)
        work_package_id = request.POST.get('work_package_id')
        if 'view_report' in request.POST and work_package_id:
            return redirect('work_packagechildren_report', work_package_id)
        elif 'pdf_button' in request.POST and work_package_id:
            return redirect('work_packagechildren_report_pdf', work_package_id)
        elif 'excel_button' in request.POST and work_package_id:
            return redirect('work_packagechildren_report_excel', work_package_id)
        # elif 'view_measurements' in request.POST and work_package_id:
        #     return redirect('work_package_report_measurement', work_package_id)
        # elif 'view_measurements_pdf' in request.POST and work_package_id:
        #     return redirect('work_package_report_measurement_pdf', work_package_id)
        # elif 'view_measurements_csv' in request.POST and work_package_id:
        #     return redirect('work_package_report_measurement_csv', work_package_id)
        # elif 'view_measurements_xlsx' in request.POST and work_package_id:
        #     return redirect('work_package_report_measurement_xlsx', work_package_id)
        else:
            messages.error(request,"You have to choose a WorkPackage")
    else:
        form = WPIdForm()
   
    return render(request, 'reports/work_packagechildren_assign.html', {
        'form': form,
        'project_id': project_id, 
        'project_description': project_description,
        'wps': wps
        })

def get_children(work_package_id):
    work_package = WorkPackage.objects.get(id=work_package_id)
    wps = WorkPackage.objects.all().order_by('name')
    wp_digits = []
    children = []

    for digit in work_package.name:
        if digit != '0':
            wp_digits.append(digit)

    prefix = ''.join(wp_digits)

    for wp in wps:
        if wp.name.startswith(prefix):
            children.append(wp.name)
    del children[0]

    return children


def prepare_work_package_context(work_package_id, user):
    active_project = ActiveProject.objects.get(user=user)
    project_id = active_project.project_id
    project = Project.objects.get(id=project_id)
    project_description = project.description

    children = get_children(work_package_id)
    wps_children = WorkPackage.objects.filter(name__in=children)

    work_package = WorkPackage.objects.get(id=work_package_id)
    work_package_description = work_package.description
    work_package_desctiption_code = list(work_package_description)
    work_package_desctiption_code_fp = work_package_desctiption_code[:8]
    work_package_desctiption_code_sp = work_package_desctiption_code[10:]
    first_part = ''.join(work_package_desctiption_code_fp)
    second_part = ''.join(work_package_desctiption_code_sp)

    letters = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    
    first_children = WorkPackage.objects.filter(parent=work_package_id).order_by('name')
    
    work_package_data = []
    for wp in wps_children:
        works = Work.objects.filter(work_package=wp).select_related('task')
        task_ids = list(works.values_list('task_id', flat=True))
        tasks_retrieved = Task.objects.filter(id__in=task_ids)
        work_package_data.append({
            'work_package': wp,
            'works': works,
            'tasks': tasks_retrieved,
        })
    
    work_packages = WorkPackage.objects.all().order_by('name')
    chapter_totals = calculate_chapter_totals(work_packages)
    
    first_children_names = [fc.name for fc in first_children]
    
    first_children_budget = {fc: chapter_totals[fc] for fc in first_children_names}
    
    work_package_budget = chapter_totals[work_package.name]

    second_children_dict = {}
    third_children_dict = {}
    third_children_tasks = {}

    for fc in first_children:
        second_children = WorkPackage.objects.filter(parent=fc.id).order_by('name')
        second_children_dict[fc.name] = second_children
        for sc in second_children:
            third_children = WorkPackage.objects.filter(parent=sc.id).order_by('name')
            for third in third_children:
                third_children_dict[sc.name] = third.name, third.description
                works = Work.objects.filter(work_package=third.id).select_related('task')
                task_ids = list(Work.objects.filter(work_package=third.id).values_list('task_id', flat=False))
                task_ids = [task_id[0] for task_id in task_ids]
                tasks_retrieved = sorted(Task.objects.filter(id__in=task_ids), key=lambda x: task_ids.index(x.id))
                third_children_tasks[third.name] = works, tasks_retrieved
                
    workpackage_dict = {}
    for workpackage_key, (work_queryset, task_list) in third_children_tasks.items():
        task_details = []
        for task in task_list:
            work_objects = work_queryset.filter(task=task)
            quantity = 0
            amount = 0
            if work_objects.exists():
                first_work = work_objects.first()
                quantity = first_work.quantity
                amount = first_work.work_amount
            task_details.append({
                'name': task.name,
                'unit': task.unit,
                'quantity': "{:,.2f}".format(quantity),
                'price': "{:,.2f}".format(task.price),
                'amount': "{:,.2f}".format(amount),
            })
            workpackage_dict[workpackage_key] = task_details

    return {
        'project_id': project_id,
        'project_description': project_description,
        'work_package_id': work_package_id,
        'work_package_name': work_package.name,
        'work_package_description': work_package.description,
        'work_package_data': work_package_data,
        'first_part': first_part,
        'second_part': second_part,
        'letters': letters,
        'children': children,
        'first_children': first_children,
        'first_children_names': first_children_names,
        'first_children_budget': first_children_budget,
        'work_package_budget': work_package_budget,
        'second_children_dict': second_children_dict,
        'third_children_dict': third_children_dict,
        'workpackage_dict': workpackage_dict,
    }


def work_packagechildren_report(request, work_package_id):
    # Prepare the context data using the helper function
    context = prepare_work_package_context(work_package_id, request.user)

    # Render the HTML template with the context data
    return render(request, 'reports/work_packagechildren_report.html', context)



def work_packagechildren_report_pdf(request, work_package_id):
    # Prepare the context data using the helper function
    context = prepare_work_package_context(work_package_id, request.user)

    # Render the HTML template with the context data
    html_string = render(request, 'reports/work_packagechildren_report.html', context).content

    # Create a PDF from the rendered HTML
    pdf = HTML(string=html_string).write_pdf()

    # Create an HTTP response with the PDF
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="work_package_report_{work_package_id}.pdf"'

    return response

# def work_packagechildren_report_excel(request, work_package_id):
#     # Define the maximum number of rows per page
#     MAX_ROWS_PER_PAGE = 43

#     def add_blank_line_with_borders(sheet):
#         # Add a blank line
#         sheet.append(["", "", "", ""])
        
#         # Set lateral borders for the blank line
#         last_row = sheet.max_row  # Get the last row number
#         for col_index in range(1, 5):  # Loop through columns A to D (1 to 4)
#             cell = sheet.cell(row=last_row, column=col_index)
#             cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))  # Apply thin border only to left and right    

#     def add_blank_line_with_borders_divisions(sheet):
#         # Add a blank line
#         sheet.append(["", "", "", "", "", "", ""])
        
#         # Set lateral borders for the blank line
#         last_row = sheet.max_row  # Get the last row number
#         for col_index in range(1, 8):  # Loop through columns A to D (1 to 4)
#             cell = sheet.cell(row=last_row, column=col_index)
#             cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))  # Apply thin border only to left and right    

#     def add_column_borders(sheet):
#         for col in range(1, 8):  # This will loop through columns 1 to 7
#                 cell = sheet.cell(row=start_row, column=col)  # Get the actual cell object
#                 cell.border = children_border  # Apply the border to the cell

#     def get_first_part(string, int):
#         string_list = []
#         for s in string[:int]:
#             string_list.append(s)
#         first_part = ''.join(string_list)
#         return first_part


#     def get_middle_part(string, int1, int2):
#         string_list = []
#         for s in string[int1:int2]:
#             string_list.append(s)
#         middle_part = ''.join(string_list)
#         return middle_part
    
#     def get_second_part(string, int):
#         string_list = []
#         for s in string[int:]:
#             string_list.append(s)
#         second_part = ''.join(string_list)
#         return second_part
    
#     # Function to get letters
#     def get_letter(index):
#         letters = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
#         if index > 25:
#             return 'N/A'  # Return 'N/A' for indices above 25
#         else: 
#             return letters[index]

#     # Prepare the context data using the helper function
#     context = prepare_work_package_context(work_package_id, request.user)

#     # Create a new Excel workbook and select the active worksheet
#     workbook = Workbook()
#     sheet = workbook.active
#     sheet.title = f"SUM - {context['first_part']}"

#     # Define styles for the headers############################################################################33
#     header_font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Black font
#     header_fill = PatternFill(start_color="FCD5B5", end_color="FCD5B5", fill_type="solid")  # Background color
#     center_alignment = Alignment(horizontal='center', vertical='center')
#     left_alignment = Alignment(horizontal='left', vertical='center')

#     # Define border styles
#     thin_border = Border(left=Side(style='thin'), 
#                         right=Side(style='thin'), 
#                         top=Side(style='thin'), 
#                         bottom=Side(style='thin'))

#     # Add the header
#     headers = ["ITEM", "SPECS", "DESCRIPTION", "AMOUNT"]
#     sheet.append(headers)

#     # Set the height of the header row (1.06 in points is approximately 15.9 in Excel)
#     sheet.row_dimensions[1].height = 30.5  # Row index is 2 because Excel is 1-indexed

#     # Style the header row
#     for cell in sheet[1]:  # Access the first row (header row)
#         cell.font = header_font
#         cell.fill = header_fill
#         cell.alignment = center_alignment
#         cell.border = thin_border 

#     # Apply double border to the bottom of the header row
#     for col_index in range(1, len(headers) + 1):  # Loop through the number of headers
#         cell = sheet.cell(row=1, column=col_index)
#         cell.border = Border(top=cell.border.top, 
#                             left=cell.border.left, 
#                             right=cell.border.right, 
#                             bottom=Side(style='double'))  # Set double border for the bottom

    
#     ########################### END of Header , starts SubHeader ###############################

#     # Add the subheader
#     subheader_first_part = context['first_part']  # Use context for first part
#     subheader_second_part = context['second_part']  # Use context for second part
    

#     # Style the subheader row
#     sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)  # Merge the first cell for first_part
#     cell_first_part = sheet.cell(row=2, column=1)
#     cell_first_part.font = Font(name='Helvetica', bold=True, size=10, color="000000", underline="single")  # Underlined font
#     cell_first_part.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")  # Background color
#     cell_first_part.alignment = center_alignment
#     cell_first_part.border = thin_border
#     cell_first_part.value = subheader_first_part 

#     # Merge cells for the second part of the subheader
#     sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)  # Merge last two cells for second_part
#     cell_second_part = sheet.cell(row=2, column=3)
#     cell_second_part.value = subheader_second_part  # Set the combined value for second_part
#     cell_second_part.font = Font(name='Helvetica', bold=True, size=10, color="000000", underline="single")  # Underlined font
#     cell_second_part.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")  # Background color
#     cell_second_part.alignment = left_alignment
#     cell_second_part.border = thin_border

#     # Set the height of the subheader row (optional)
#     sheet.row_dimensions[2].height = 32  # Row index is 2 for the subheader row

#     # Set the column widths
#     column_widths = [8, 12, 60, 25]  # Widths for columns A, B, C, D
#     for col_index, width in enumerate(column_widths, start=1):  # start=1 for A=1, B=2, C=3, D=4
#         sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = width
#     ############################################### End of Subheader ######################################################33
    
    
#     add_blank_line_with_borders(sheet)

#     # Add summary section
#     sheet.append(["", "", "SUMMARY", ""])
#     # Set border for the blank line
#     last_row = sheet.max_row  # Get the last row number
#     for col_index in range(1, 5):  # Loop through columns A to D (1 to 4)
#         cell = sheet.cell(row=last_row, column=col_index)
#         cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))  # Apply thin border only to left and right
#     # Apply bold font to the "SUMMARY" cell
#     summary_cell = sheet.cell(row=last_row, column=3)  # Column C is the third column
#     summary_cell.font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Set font to bold

#     add_blank_line_with_borders(sheet)

#     for index, first_child in enumerate(context.get('first_children', [])):
#         sheet.append([get_letter(index), "",  first_child.description,  context['first_children_budget'].get(first_child.name, "")])
#         # Set border for the blank line
#         last_row = sheet.max_row  # Get the last row number
#         for col_index in range(1, 5):  # Loop through columns A to D (1 to 4)
#             cell = sheet.cell(row=last_row, column=col_index)
#             cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))  # Apply thin border only to left and right
#             # Apply font style to the Division lines
#             summary_cell = sheet.cell(row=last_row, column=col_index)  
#             summary_cell.font = Font(name='Helvetica', bold=False, size=9, color="000000")  # Set font to bold
#             cell = sheet.cell(row=last_row, column=1)
#             cell.alignment = Alignment(horizontal='center', vertical='center')

#     # Get the current number of filled rows
#     current_filled_rows = sheet.max_row
    
#     # Calculate remaining lines
#     remaining_lines = MAX_ROWS_PER_PAGE - (current_filled_rows % MAX_ROWS_PER_PAGE)
    

      
#     # Calculate how many blank lines to add
#     lines_to_add = remaining_lines + 1 
#     for _ in range(lines_to_add):
#         add_blank_line_with_borders(sheet)


#     # Total row
#     sheet.append(["", "", f"{context['first_part']} TOTAL CARRIED TO MAIN SCHEDULE", context['work_package_budget']])
#     last_row = sheet.max_row  # Get the last row number
#     # Set the height of the total row
#     sheet.row_dimensions[last_row].height = 30.5
#     for col_index in range(1, 5):  # Loop through columns A to D (1 to 4)
#         cell = sheet.cell(row=last_row, column=col_index)
#         cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='medium'),
#             bottom=Side(style='medium'))  # Apply thin border only to left and right and thick to top and bottom
#         cell.alignment = Alignment(horizontal='center', vertical='center')
#         # Apply font style to the Division lines
#         summary_cell = sheet.cell(row=last_row, column=col_index)  
#         summary_cell.font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Set font to bold
#     ############################################################### DIVISION SHEETS ###############################################
#     ###############################################################################################################################    
#     for first_child in context['first_children']:
#         div_number = get_middle_part(first_child.description, 9, 11)
#         sheet = workbook.create_sheet(title=f"D{div_number}")

#         division_headers = ["ITEM", "SPECS", "DESCRIPTION", "UNIT", "QTY", "RATE", "AMOUNT"]
#         sheet.append(division_headers)
#         last_row = sheet.max_row
#         sheet.row_dimensions[last_row].height = 30.5

#         column_widths = [6, 10, 35, 6, 15, 15, 20]

#         for col_index in range(1, 8):  # Loop through columns A to G (1 to 7)
#             sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = column_widths[col_index - 1]
#             cell = sheet.cell(row=last_row, column=col_index)
#             cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
#                 bottom=Side(style='double'))  # Apply thin border only to left and right and thick to top and bottom
#             cell.font = header_font
#             cell.alignment = center_alignment
#             cell.fill = header_fill

#          # Add the subheader ################################################################
#         subheader_first_part = context['first_part']  # Use context for first part
#         subheader_second_part = context['second_part']  # Use context for second part
        

#         # Style the subheader row
#         sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)  # Merge the first cell for first_part
#         cell_first_part = sheet.cell(row=2, column=1)
#         cell_first_part.font = Font(name='Helvetica', bold=True, size=10, color="000000", underline="single")  # Underlined font
#         cell_first_part.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")  # Background color
#         cell_first_part.alignment = center_alignment
#         cell_first_part.border = thin_border
#         cell_first_part.value = subheader_first_part 

#         # Merge cells for the second part of the subheader
#         sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=6)  # Merge last two cells for second_part
#         cell_second_part = sheet.cell(row=2, column=3)
#         cell_second_part.value = subheader_second_part  # Set the combined value for second_part
#         cell_second_part.font = Font(name='Helvetica', bold=True, size=10, color="000000", underline="single")  # Underlined font
#         cell_second_part.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")  # Background color
#         cell_second_part.alignment = left_alignment
#         cell_second_part.border = Border(left=Side(style='thin'), right=Side(style='thin'),  # Add thin right border
#                       top=Side(style='double'), bottom=Side(style='thin'))
#         last_row = sheet.max_row
#         cell = sheet.cell(row=last_row, column=7)
#         cell.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")
#         cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),  # Add thin right border
#                       top=Side(style='double'), bottom=Side(style='thin'))

#         # Set the height of the subheader row (optional)
#         sheet.row_dimensions[2].height = 32  # Row index is 2 for the subheader row

#         # End of subheader###############################################################
#         # Add division data
#         last_row = sheet.max_row + 1 
#         # Set the content of the third column in the last row
#         division_cell = sheet.cell(row=last_row, column=3)
#         division_cell.value = first_child.description.upper()  # Set the description in the third column

#         # Apply thin borders to all columns for the last row
#         for col_index in range(1, 8):  # Loop through columns A to G (1 to 7)
#             cell = sheet.cell(row=last_row, column=col_index)
#             cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
#                                 top=Side(style='thin'), bottom=Side(style='thin'))  # Apply thin borders on all sides

#         # Set the font style for the content in the third column
#         division_cell.font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Set font to bold
        
#         add_blank_line_with_borders_divisions(sheet)
#         # styles
#         children_font_bold = Font(name='Helvetica', bold=True, size=10, color="000000")  # Black font
#         children_font = Font(name='Helvetica', bold=False, size=9, color="000000")  # Black font
#         children_font_italic_underline = Font(name='Helvetica', italic=True, underline='single', size=10, color="000000")
#         children_alignment_center = Alignment(horizontal='center', vertical='center')
#         children_alignment_left = Alignment(horizontal='left', vertical='center')
#         children_border = Border(left=Side(style='thin'), right=Side(style='thin'))

#         second_children_dictionary = context['second_children_dict']
#         second_children_object = second_children_dictionary[first_child.name]

#         # Start writing from the last row + 1 to avoid overwriting
#         start_row = sheet.max_row + 1

#         page_value = 0

#         for second in second_children_object:
#             first_part = get_first_part(second.description, 8)
#             second_part = get_second_part(second.description, 11)
            
#             # Set the height of the total row
#             cell = sheet.cell(row=start_row, column=2)
#             cell.border = children_border
#             cell.alignment = children_alignment_center
#             cell.font = children_font_bold
#             cell.value = first_part
#             cell = sheet.cell(row=start_row, column=3)
#             cell.alignment = children_alignment_left
#             cell.font = children_font_bold
#             cell.value = second_part
#             add_column_borders(sheet)

#             # Move to the next row for the next object
#             start_row += 1

#             third_children_dictionary = context['third_children_dict']
#             third_children = third_children_dictionary[second.name]

#             cell = sheet.cell(row=start_row, column=3)
#             cell.alignment = children_alignment_left
#             cell.font = children_font_italic_underline
#             cell.value = third_children[1]
#             add_column_borders(sheet)

#             start_row += 1

#             workpackage_dict = context['workpackage_dict']
#             tasks = workpackage_dict[third_children[0]]

#             for task in tasks:
#                 cell = sheet.cell(row=start_row, column=3)
#                 cell.alignment = children_alignment_left
#                 cell.font = children_font
#                 cell.value = task['name']
                

#                 cell = sheet.cell(row=start_row, column=4)
#                 cell.alignment = children_alignment_center
#                 cell.font = children_font
#                 cell.value = task['unit']
                

#                 cell = sheet.cell(row=start_row, column=5)
#                 cell.alignment = children_alignment_center
#                 cell.font = children_font
#                 cell.value = task['quantity']

#                 cell = sheet.cell(row=start_row, column=6)
#                 cell.alignment = children_alignment_center
#                 cell.font = children_font
#                 cell.value = task['price']

#                 cell = sheet.cell(row=start_row, column=7)
#                 cell.alignment = children_alignment_center
#                 cell.font = children_font
#                 cell.value = task['amount']
#                 amount_str = task['amount'].strip().replace(',', '')
#                 page_value += float(amount_str)
                
#                 add_column_borders(sheet)
#                 start_row += 1

                
#         #################################Summary at the end of the page ###################################
#         # Calculate how many blank lines to add
#         current_filled_rows = sheet.max_row
        
#         print("current filled rows: ", current_filled_rows)
#         # Calculate remaining lines
#         remaining_lines = MAX_ROWS_PER_PAGE - (current_filled_rows)
        
#         lines_to_add = remaining_lines + 1 

#         print("lines to add: ", lines_to_add)

        
#         for _ in range(lines_to_add):
#             add_blank_line_with_borders_divisions(sheet)


#         # Total row
#         sheet.append(["", "", f"PAGE - TOTAL CARRIED TO {first_child.description.upper()}"])
#         last_row = sheet.max_row  # Get the last row number
#         # Set the height of the total row
#         sheet.row_dimensions[last_row].height = 30.5
#         for col_index in range(1, 8):  # Loop through columns A to G (1 to 7)
#             cell = sheet.cell(row=last_row, column=col_index)
#             cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='medium'),
#                 bottom=Side(style='medium'))  # Apply thin border only to left and right and thick to top and bottom
#             cell.alignment = Alignment(horizontal='left', vertical='center')
#             # Apply font style to the Division lines
#         summary_cell = sheet.cell(row=last_row, column=7)  
#         summary_cell.font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Set font to bold
#         summary_cell.alignment = children_alignment_center
#         summary_cell.value = "{:,.2f}".format(page_value)
        
    
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = f'attachment; filename="{context["first_part"]}.xlsx"'
#     workbook.save(response)

#     return response

def work_packagechildren_report_excel(request, work_package_id):
    # Define the maximum number of rows per page
    MAX_ROWS_PER_PAGE = 43

    def add_blank_line_with_borders(sheet):
        # Add a blank line
        sheet.append(["", "", "", ""])
        
        # Set lateral borders for the blank line
        last_row = sheet.max_row  # Get the last row number
        for col_index in range(1, 5):  # Loop through columns A to D (1 to 4)
            cell = sheet.cell(row=last_row, column=col_index)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))  # Apply thin border only to left and right    

    def add_blank_line_with_borders_divisions(sheet):
        # Add a blank line
        sheet.append(["", "", "", "", "", "", ""])
        
        # Set lateral borders for the blank line
        last_row = sheet.max_row  # Get the last row number
        for col_index in range(1, 8):  # Loop through columns A to D (1 to 4)
            cell = sheet.cell(row=last_row, column=col_index)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))  # Apply thin border only to left and right    

    def add_column_borders(sheet):
        for col in range(1, 8):  # This will loop through columns 1 to 7
                cell = sheet.cell(row=start_row, column=col)  # Get the actual cell object
                cell.border = children_border  # Apply the border to the cell

    def get_first_part(string, int):
        string_list = []
        for s in string[:int]:
            string_list.append(s)
        first_part = ''.join(string_list)
        return first_part


    def get_middle_part(string, int1, int2):
        string_list = []
        for s in string[int1:int2]:
            string_list.append(s)
        middle_part = ''.join(string_list)
        return middle_part
    
    def get_second_part(string, int):
        string_list = []
        for s in string[int:]:
            string_list.append(s)
        second_part = ''.join(string_list)
        return second_part
    
    # Function to get letters
    def get_letter(index):
        letters = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
        if index > 25:
            return 'N/A'  # Return 'N/A' for indices above 25
        else: 
            return letters[index]

    # Prepare the context data using the helper function
    context = prepare_work_package_context(work_package_id, request.user)

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"SUM - {context['first_part']}"

    # Define styles for the headers############################################################################33
    header_font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Black font
    header_fill = PatternFill(start_color="FCD5B5", end_color="FCD5B5", fill_type="solid")  # Background color
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')

    # Define border styles
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # Add the header
    headers = ["ITEM", "SPECS", "DESCRIPTION", "AMOUNT"]
    sheet.append(headers)

    # Set the height of the header row (1.06 in points is approximately 15.9 in Excel)
    sheet.row_dimensions[1].height = 30.5  # Row index is 2 because Excel is 1-indexed

    # Style the header row
    for cell in sheet[1]:  # Access the first row (header row)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border 

    # Apply double border to the bottom of the header row
    for col_index in range(1, len(headers) + 1):  # Loop through the number of headers
        cell = sheet.cell(row=1, column=col_index)
        cell.border = Border(top=cell.border.top, 
                            left=cell.border.left, 
                            right=cell.border.right, 
                            bottom=Side(style='double'))  # Set double border for the bottom

    
    ########################### END of Header , starts SubHeader ###############################

    # Add the subheader
    subheader_first_part = context['first_part']  # Use context for first part
    subheader_second_part = context['second_part']  # Use context for second part
    

    # Style the subheader row
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)  # Merge the first cell for first_part
    cell_first_part = sheet.cell(row=2, column=1)
    cell_first_part.font = Font(name='Helvetica', bold=True, size=10, color="000000", underline="single")  # Underlined font
    cell_first_part.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")  # Background color
    cell_first_part.alignment = center_alignment
    cell_first_part.border = thin_border
    cell_first_part.value = subheader_first_part 

    # Merge cells for the second part of the subheader
    sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)  # Merge last two cells for second_part
    cell_second_part = sheet.cell(row=2, column=3)
    cell_second_part.value = subheader_second_part  # Set the combined value for second_part
    cell_second_part.font = Font(name='Helvetica', bold=True, size=10, color="000000", underline="single")  # Underlined font
    cell_second_part.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")  # Background color
    cell_second_part.alignment = left_alignment
    cell_second_part.border = thin_border

    # Set the height of the subheader row (optional)
    sheet.row_dimensions[2].height = 32  # Row index is 2 for the subheader row

    # Set the column widths
    column_widths = [8, 12, 60, 25]  # Widths for columns A, B, C, D
    for col_index, width in enumerate(column_widths, start=1):  # start=1 for A=1, B=2, C=3, D=4
        sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = width
    ############################################### End of Subheader ######################################################33
    
    
    add_blank_line_with_borders(sheet)

    # Add summary section
    sheet.append(["", "", "SUMMARY", ""])
    # Set border for the blank line
    last_row = sheet.max_row  # Get the last row number
    for col_index in range(1, 5):  # Loop through columns A to D (1 to 4)
        cell = sheet.cell(row=last_row, column=col_index)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))  # Apply thin border only to left and right
    # Apply bold font to the "SUMMARY" cell
    summary_cell = sheet.cell(row=last_row, column=3)  # Column C is the third column
    summary_cell.font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Set font to bold

    add_blank_line_with_borders(sheet)

    for index, first_child in enumerate(context.get('first_children', [])):
        sheet.append([get_letter(index), "",  first_child.description,  context['first_children_budget'].get(first_child.name, "")])
        # Set border for the blank line
        last_row = sheet.max_row  # Get the last row number
        for col_index in range(1, 5):  # Loop through columns A to D (1 to 4)
            cell = sheet.cell(row=last_row, column=col_index)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))  # Apply thin border only to left and right
            # Apply font style to the Division lines
            summary_cell = sheet.cell(row=last_row, column=col_index)  
            summary_cell.font = Font(name='Helvetica', bold=False, size=9, color="000000")  # Set font to bold
            cell = sheet.cell(row=last_row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Get the current number of filled rows
    current_filled_rows = sheet.max_row
    
    # Calculate remaining lines
    remaining_lines = MAX_ROWS_PER_PAGE - (current_filled_rows % MAX_ROWS_PER_PAGE)
    

      
    # Calculate how many blank lines to add
    lines_to_add = remaining_lines + 1 
    for _ in range(lines_to_add):
        add_blank_line_with_borders(sheet)


    # Total row
    sheet.append(["", "", f"{context['first_part']} TOTAL CARRIED TO MAIN SCHEDULE", context['work_package_budget']])
    last_row = sheet.max_row  # Get the last row number
    # Set the height of the total row
    sheet.row_dimensions[last_row].height = 30.5
    for col_index in range(1, 5):  # Loop through columns A to D (1 to 4)
        cell = sheet.cell(row=last_row, column=col_index)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='medium'),
            bottom=Side(style='medium'))  # Apply thin border only to left and right and thick to top and bottom
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Apply font style to the Division lines
        summary_cell = sheet.cell(row=last_row, column=col_index)  
        summary_cell.font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Set font to bold
    ############################################################### DIVISION SHEETS ###############################################
    ###############################################################################################################################
    page_totals = []
    current_page_total = 0 
       
    for first_child in context['first_children']:
        div_number = get_middle_part(first_child.description, 9, 11)
        sheet = workbook.create_sheet(title=f"D{div_number}")

        division_headers = ["ITEM", "SPECS", "DESCRIPTION", "UNIT", "QTY", "RATE", "AMOUNT"]
        sheet.append(division_headers)
        last_row = sheet.max_row
        sheet.row_dimensions[last_row].height = 30.5

        column_widths = [6, 10, 35, 6, 15, 15, 20]

        for col_index in range(1, 8):  # Loop through columns A to G (1 to 7)
            sheet.column_dimensions[sheet.cell(row=1, column=col_index).column_letter].width = column_widths[col_index - 1]
            cell = sheet.cell(row=last_row, column=col_index)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                bottom=Side(style='double'))  # Apply thin border only to left and right and thick to top and bottom
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill

         # Add the subheader ################################################################
        subheader_first_part = context['first_part']  # Use context for first part
        subheader_second_part = context['second_part']  # Use context for second part
        

        # Style the subheader row
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)  # Merge the first cell for first_part
        cell_first_part = sheet.cell(row=2, column=1)
        cell_first_part.font = Font(name='Helvetica', bold=True, size=10, color="000000", underline="single")  # Underlined font
        cell_first_part.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")  # Background color
        cell_first_part.alignment = center_alignment
        cell_first_part.border = thin_border
        cell_first_part.value = subheader_first_part 

        # Merge cells for the second part of the subheader
        sheet.merge_cells(start_row=2, start_column=3, end_row=2, end_column=6)  # Merge last two cells for second_part
        cell_second_part = sheet.cell(row=2, column=3)
        cell_second_part.value = subheader_second_part  # Set the combined value for second_part
        cell_second_part.font = Font(name='Helvetica', bold=True, size=10, color="000000", underline="single")  # Underlined font
        cell_second_part.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")  # Background color
        cell_second_part.alignment = left_alignment
        cell_second_part.border = Border(left=Side(style='thin'), right=Side(style='thin'),  # Add thin right border
                      top=Side(style='double'), bottom=Side(style='thin'))
        last_row = sheet.max_row
        cell = sheet.cell(row=last_row, column=7)
        cell.fill = PatternFill(start_color="FDEADA", end_color="FDEADA", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),  # Add thin right border
                      top=Side(style='double'), bottom=Side(style='thin'))

        # Set the height of the subheader row (optional)
        sheet.row_dimensions[2].height = 32  # Row index is 2 for the subheader row

        # End of subheader###############################################################
        # Add division data
        last_row = sheet.max_row + 1 
        # Set the content of the third column in the last row
        division_cell = sheet.cell(row=last_row, column=3)
        division_cell.value = first_child.description.upper()  # Set the description in the third column

        # Apply thin borders to all columns for the last row
        for col_index in range(1, 8):  # Loop through columns A to G (1 to 7)
            cell = sheet.cell(row=last_row, column=col_index)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                top=Side(style='thin'), bottom=Side(style='thin'))  # Apply thin borders on all sides

        # Set the font style for the content in the third column
        division_cell.font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Set font to bold
        
        add_blank_line_with_borders_divisions(sheet)
        # styles
        children_font_bold = Font(name='Helvetica', bold=True, size=10, color="000000")  # Black font
        children_font = Font(name='Helvetica', bold=False, size=9, color="000000")  # Black font
        children_font_italic_underline = Font(name='Helvetica', italic=True, underline='single', size=10, color="000000")
        children_alignment_center = Alignment(horizontal='center', vertical='center')
        children_alignment_left = Alignment(horizontal='left', vertical='center')
        children_border = Border(left=Side(style='thin'), right=Side(style='thin'))

        second_children_dictionary = context['second_children_dict']
        second_children_object = second_children_dictionary[first_child.name]

        # Start writing from the last row + 1 to avoid overwriting
        start_row = sheet.max_row + 1

        page_value = 0

        for second in second_children_object:
            first_part = get_first_part(second.description, 8)
            second_part = get_second_part(second.description, 11)

            if start_row >= MAX_ROWS_PER_PAGE:
                # Add the current page total to the page_totals list
                page_totals.append(current_page_total)

                add_blank_line_with_borders_divisions(sheet)

                # Reset start row for the next page section
                start_row += 1  # Move down for the next page section
                current_page_total = 0  # Reset page total for new page section
            
            
            # Set the height of the total row
            cell = sheet.cell(row=start_row, column=2)
            cell.border = children_border
            cell.alignment = children_alignment_center
            cell.font = children_font_bold
            cell.value = first_part
            cell = sheet.cell(row=start_row, column=3)
            cell.alignment = children_alignment_left
            cell.font = children_font_bold
            cell.value = second_part
            add_column_borders(sheet)

            # Move to the next row for the next object
            start_row += 1

            third_children_dictionary = context['third_children_dict']
            third_children = third_children_dictionary[second.name]

            cell = sheet.cell(row=start_row, column=3)
            cell.alignment = children_alignment_left
            cell.font = children_font_italic_underline
            cell.value = third_children[1]
            add_column_borders(sheet)

            start_row += 1

            workpackage_dict = context['workpackage_dict']
            tasks = workpackage_dict[third_children[0]]

            for task in tasks:
                cell = sheet.cell(row=start_row, column=3)
                cell.alignment = children_alignment_left
                cell.font = children_font
                cell.value = task['name']
                

                cell = sheet.cell(row=start_row, column=4)
                cell.alignment = children_alignment_center
                cell.font = children_font
                cell.value = task['unit']
                

                cell = sheet.cell(row=start_row, column=5)
                cell.alignment = children_alignment_center
                cell.font = children_font
                cell.value = task['quantity']

                cell = sheet.cell(row=start_row, column=6)
                cell.alignment = children_alignment_center
                cell.font = children_font
                cell.value = task['price']

                cell = sheet.cell(row=start_row, column=7)
                cell.alignment = children_alignment_center
                cell.font = children_font
                cell.value = task['amount']
                amount_str = task['amount'].strip().replace(',', '')
                page_value += float(amount_str)
                
                add_column_borders(sheet)
                start_row += 1

        # After processing all second children, check the remaining page total
        if current_page_total > 0:
            page_totals.append(current_page_total)

                
        #################################Summary at the end of the page ###################################
        # Calculate how many blank lines to add
        current_filled_rows = sheet.max_row
        
        print("current filled rows: ", current_filled_rows)
        # Calculate remaining lines
        remaining_lines = MAX_ROWS_PER_PAGE - (current_filled_rows)
        
        lines_to_add = remaining_lines + 1 

        print("lines to add: ", lines_to_add)

        
        for _ in range(lines_to_add):
            add_blank_line_with_borders_divisions(sheet)


        # Total row
        sheet.append(["", "", f"PAGE - TOTAL CARRIED TO {first_child.description.upper()}"])
        last_row = sheet.max_row  # Get the last row number
        # Set the height of the total row
        sheet.row_dimensions[last_row].height = 30.5
        for col_index in range(1, 8):  # Loop through columns A to G (1 to 7)
            cell = sheet.cell(row=last_row, column=col_index)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='medium'),
                bottom=Side(style='medium'))  # Apply thin border only to left and right and thick to top and bottom
            cell.alignment = Alignment(horizontal='left', vertical='center')
            # Apply font style to the Division lines
        summary_cell = sheet.cell(row=last_row, column=7)  
        summary_cell.font = Font(name='Helvetica', bold=True, size=10, color="000000")  # Set font to bold
        summary_cell.alignment = children_alignment_center
        summary_cell.value = "{:,.2f}".format(page_value)
        
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{context["first_part"]}.xlsx"'
    workbook.save(response)

    return response